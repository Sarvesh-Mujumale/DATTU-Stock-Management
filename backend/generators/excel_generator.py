"""
Excel Generator Module
======================
Generates standardized Excel files from extracted data.

Features:
- Fixed column structure for consistency
- Status/Error column for validation results
- Optional validation notes sheet
- Inventory analysis with surplus/deficit (NEW)
- 100% in-memory generation using BytesIO

No data is written to disk. Buffer is cleared after use.
"""

import io
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from extraction import ExtractedData
from validation import ValidationResult


class ExcelGenerator:
    """
    In-memory Excel generator for processed documents.
    
    Creates standardized Excel files with:
    - Summary sheet with extracted data
    - Line items sheet (if items present)
    - Validation notes sheet
    
    All operations are performed in BytesIO buffers.
    No data is written to disk.
    """
    
    # Standard column widths
    COLUMN_WIDTHS = {
        'A': 18,  # Field names
        'B': 35,  # Values
        'C': 15,  # Status
        'D': 40,  # Notes
    }
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def generate(
        self,
        data: ExtractedData,
        validation: ValidationResult,
        original_filename: str = ""
    ) -> bytes:
        """
        Generate Excel file from extracted data.
        
        Args:
            data: ExtractedData object with extracted fields
            validation: ValidationResult with errors and warnings
            original_filename: Original uploaded filename for reference
            
        Returns:
            Bytes of the generated Excel file
        """
        # Create workbook in memory
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, data, validation, original_filename)
        self._create_line_items_sheet(wb, data)
        self._create_validation_sheet(wb, data, validation)
        
        # Remove default empty sheet if we created others
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
        
        # Write to BytesIO buffer
        buffer = io.BytesIO()
        try:
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        finally:
            # Clear buffer
            buffer.close()
            wb.close()
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        data: ExtractedData,
        validation: ValidationResult,
        original_filename: str
    ):
        """
        Create the main summary sheet with extracted data.
        """
        ws = wb.create_sheet("Summary", 0)
        
        # Set column widths
        for col, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width
        
        # Header row
        headers = ["Field", "Value", "Status", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Determine overall status
        if validation.is_valid and not validation.warnings:
            overall_status = "Valid"
            status_fill = self.SUCCESS_FILL
        elif validation.is_valid:
            overall_status = "Valid (with warnings)"
            status_fill = self.WARNING_FILL
        else:
            overall_status = "Invalid"
            status_fill = self.ERROR_FILL
        
        # Data rows with pricing and GST
        rows = [
            ("Source File", original_filename or "Uploaded Document", "", ""),
            ("Processing Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""),
            ("Overall Status", overall_status, overall_status, f"{len(validation.errors)} errors, {len(validation.warnings)} warnings"),
            ("", "", "", ""),  # Empty row
            ("Invoice/Bill Number", data.invoice_number or "Not found", "Warning" if not data.invoice_number else "OK", ""),
            ("Date", data.date or "Not found", "Warning" if not data.date else "OK", ""),
            ("Vendor/Supplier", data.vendor_name or "Not found", "Warning" if not data.vendor_name else "OK", ""),
            ("", "", "", ""),  # Empty row
            ("Total Line Items", str(len(data.line_items)), "OK" if len(data.line_items) > 0 else "Warning", "Number of items extracted"),
            ("Total Quantity", str(sum(item.quantity for item in data.line_items)), "", "Sum of all item quantities"),
            ("", "", "", ""),  # Empty row
            ("Subtotal", f"{data.subtotal:.2f}" if data.subtotal > 0 else "Not found", "", "Before GST"),
            ("CGST", f"{data.cgst:.2f}" if data.cgst > 0 else "-", "", "Central GST"),
            ("SGST", f"{data.sgst:.2f}" if data.sgst > 0 else "-", "", "State GST"),
            ("IGST", f"{data.igst:.2f}" if data.igst > 0 else "-", "", "Integrated GST"),
            ("Total GST", f"{data.tax:.2f}" if data.tax > 0 else "Not found", "", "CGST + SGST or IGST"),
            ("Grand Total", f"{data.total:.2f}" if data.total > 0 else "Not found", "OK" if data.total > 0 else "Warning", "Final amount including GST"),
        ]
        
        for row_num, (field, value, status, notes) in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=field).border = self.BORDER
            ws.cell(row=row_num, column=2, value=value).border = self.BORDER
            
            status_cell = ws.cell(row=row_num, column=3, value=status)
            status_cell.border = self.BORDER
            if status == "OK":
                status_cell.fill = self.SUCCESS_FILL
            elif status == "Warning":
                status_cell.fill = self.WARNING_FILL
            elif status == "Error":
                status_cell.fill = self.ERROR_FILL
            elif "Valid" in status and "Invalid" not in status:
                status_cell.fill = self.SUCCESS_FILL if "warnings" not in status else self.WARNING_FILL
            elif "Invalid" in status:
                status_cell.fill = self.ERROR_FILL
            
            ws.cell(row=row_num, column=4, value=notes).border = self.BORDER
    
    def _create_line_items_sheet(self, wb: Workbook, data: ExtractedData):
        """
        Create sheet with line item details including prices and discount percentage.
        """
        ws = wb.create_sheet("Line Items")
        
        # Set column widths
        ws.column_dimensions['A'].width = 8    # S.No
        ws.column_dimensions['B'].width = 40   # Item Name
        ws.column_dimensions['C'].width = 10   # Quantity
        ws.column_dimensions['D'].width = 12   # Rate
        ws.column_dimensions['E'].width = 10   # Disc. %
        ws.column_dimensions['F'].width = 14   # Amount
        
        # Header row
        headers = ["S.No", "Item/Description", "Qty", "Rate", "Disc. %", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Data rows with prices and discount percentage
        if data.line_items:
            for row_num, item in enumerate(data.line_items, 2):
                ws.cell(row=row_num, column=1, value=row_num - 1).border = self.BORDER
                ws.cell(row=row_num, column=2, value=item.item_name).border = self.BORDER
                
                qty_cell = ws.cell(row=row_num, column=3, value=item.quantity)
                qty_cell.alignment = Alignment(horizontal='center')
                qty_cell.border = self.BORDER
                
                rate_cell = ws.cell(row=row_num, column=4, value=item.rate if item.rate > 0 else "")
                rate_cell.alignment = Alignment(horizontal='right')
                rate_cell.border = self.BORDER
                rate_cell.number_format = '#,##0.00'
                
                # Discount percentage column - show as "50%" format
                disc_cell = ws.cell(row=row_num, column=5, value=f"{int(item.discount_percent)}%" if item.discount_percent > 0 else "")
                disc_cell.alignment = Alignment(horizontal='center')
                disc_cell.border = self.BORDER
                
                amount_cell = ws.cell(row=row_num, column=6, value=item.amount if item.amount > 0 else "")
                amount_cell.alignment = Alignment(horizontal='right')
                amount_cell.border = self.BORDER
                amount_cell.number_format = '#,##0.00'
            
            # Total row
            total_row = len(data.line_items) + 2
            ws.cell(row=total_row, column=1).border = self.BORDER
            
            total_label = ws.cell(row=total_row, column=2, value="TOTAL")
            total_label.font = Font(bold=True)
            total_label.border = self.BORDER
            
            total_qty = sum(item.quantity for item in data.line_items)
            total_qty_cell = ws.cell(row=total_row, column=3, value=total_qty)
            total_qty_cell.font = Font(bold=True)
            total_qty_cell.alignment = Alignment(horizontal='center')
            total_qty_cell.border = self.BORDER
            
            ws.cell(row=total_row, column=4).border = self.BORDER
            
            # Empty cell for discount column in total row
            ws.cell(row=total_row, column=5).border = self.BORDER
            
            total_amount = sum(item.amount for item in data.line_items)
            total_amount_cell = ws.cell(row=total_row, column=6, value=total_amount if total_amount > 0 else "")
            total_amount_cell.font = Font(bold=True)
            total_amount_cell.alignment = Alignment(horizontal='right')
            total_amount_cell.fill = self.SUCCESS_FILL
            total_amount_cell.border = self.BORDER
            total_amount_cell.number_format = '#,##0.00'
        else:
            # No items message
            ws.cell(row=2, column=1, value="No line items extracted")
            ws.merge_cells('A2:F2')
    
    def _create_validation_sheet(
        self,
        wb: Workbook,
        data: ExtractedData,
        validation: ValidationResult
    ):
        """
        Create sheet with validation notes and extraction details.
        """
        ws = wb.create_sheet("Validation Notes")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        
        # Header
        header_cell = ws.cell(row=1, column=1, value="Type")
        header_cell.fill = self.HEADER_FILL
        header_cell.font = self.HEADER_FONT
        header_cell.border = self.BORDER
        
        header_cell = ws.cell(row=1, column=2, value="Message")
        header_cell.fill = self.HEADER_FILL
        header_cell.font = self.HEADER_FONT
        header_cell.border = self.BORDER
        
        row_num = 2
        
        # Errors
        for error in validation.errors:
            type_cell = ws.cell(row=row_num, column=1, value="ERROR")
            type_cell.fill = self.ERROR_FILL
            type_cell.border = self.BORDER
            
            ws.cell(row=row_num, column=2, value=error).border = self.BORDER
            row_num += 1
        
        # Warnings
        for warning in validation.warnings:
            type_cell = ws.cell(row=row_num, column=1, value="WARNING")
            type_cell.fill = self.WARNING_FILL
            type_cell.border = self.BORDER
            
            ws.cell(row=row_num, column=2, value=warning).border = self.BORDER
            row_num += 1
        
        # Extraction notes
        for note in data.extraction_notes:
            type_cell = ws.cell(row=row_num, column=1, value="INFO")
            type_cell.border = self.BORDER
            
            ws.cell(row=row_num, column=2, value=note).border = self.BORDER
            row_num += 1
        
        # If no notes
        if row_num == 2:
            ws.cell(row=2, column=1, value="INFO").border = self.BORDER
            ws.cell(row=2, column=2, value="No validation issues found").border = self.BORDER

    # ========================================================================
    # Analysis Report Generation (for Sales/Purchase Bills)
    # ========================================================================
    
    def generate_analysis_report(
        self,
        analysis,  # InventoryAnalysis object
        purchase_bills: List[dict],
        sales_bills: List[dict]
    ) -> bytes:
        """
        Generate comprehensive analysis Excel from purchase and sales data.
        
        Creates:
        - Inventory Summary sheet
        - Purchase Bills sheet
        - Sales Bills sheet
        - Analysis & Insights sheet
        
        Args:
            analysis: InventoryAnalysis object with calculated metrics
            purchase_bills: List of extracted purchase bill data
            sales_bills: List of extracted sales bill data
            
        Returns:
            Bytes of the generated Excel file
        """
        wb = Workbook()
        
        # Create sheets
        self._create_inventory_summary_sheet(wb, analysis)
        self._create_bills_sheet(wb, purchase_bills, "Purchase Bills", "PURCHASE")
        self._create_bills_sheet(wb, sales_bills, "Sales Bills", "SALES")
        self._create_insights_sheet(wb, analysis)
        
        # Remove default empty sheet
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
        
        # Write to buffer
        buffer = io.BytesIO()
        try:
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        finally:
            buffer.close()
            wb.close()
    
    def _create_inventory_summary_sheet(self, wb: Workbook, analysis):
        """
        Create inventory summary with quantities and values.
        """
        ws = wb.create_sheet("Stock Summary", 0)
        
        # Column widths
        ws.column_dimensions['A'].width = 35  # Item Name
        ws.column_dimensions['B'].width = 12  # Purchased Qty
        ws.column_dimensions['C'].width = 14  # Purchase Value
        ws.column_dimensions['D'].width = 10  # Sold Qty
        ws.column_dimensions['E'].width = 14  # Sales Value
        ws.column_dimensions['F'].width = 12  # Stock Balance
        ws.column_dimensions['G'].width = 12  # Status
        
        # Title
        title_cell = ws.cell(row=1, column=1, value="INVENTORY ANALYSIS REPORT")
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        
        # Generated date
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # ============ PERIOD COVERED SECTION ============
        row_num = 4
        
        # Get date ranges
        p_start, p_end = analysis.purchase_date_range
        s_start, s_end = analysis.sales_date_range
        
        # Period header
        period_cell = ws.cell(row=row_num, column=1, value="PERIOD COVERED")
        period_cell.font = Font(bold=True, size=11, color="FFFFFF")
        period_cell.fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
        period_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{row_num}:G{row_num}')
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = self.BORDER
            ws.cell(row=row_num, column=col).fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
        row_num += 1
        
        # Purchase date range
        purchase_date_text = f"{p_start} to {p_end}" if p_start and p_end else "N/A"
        if p_start == p_end and p_start:
            purchase_date_text = p_start
        
        ws.cell(row=row_num, column=1, value="Purchase Bills:").font = Font(bold=True)
        ws.cell(row=row_num, column=1).border = self.BORDER
        ws.cell(row=row_num, column=2, value=purchase_date_text).border = self.BORDER
        ws.merge_cells(f'B{row_num}:D{row_num}')
        
        # Sales date range
        sales_date_text = f"{s_start} to {s_end}" if s_start and s_end else "N/A"
        if s_start == s_end and s_start:
            sales_date_text = s_start
        
        ws.cell(row=row_num, column=5, value="Sales Bills:").font = Font(bold=True)
        ws.cell(row=row_num, column=5).border = self.BORDER
        ws.cell(row=row_num, column=6, value=sales_date_text).border = self.BORDER
        ws.merge_cells(f'F{row_num}:G{row_num}')
        for col in range(2, 8):
            ws.cell(row=row_num, column=col).border = self.BORDER
        row_num += 1
        
        # Date mismatch warning
        if analysis.date_mismatch_warning:
            warning_cell = ws.cell(row=row_num, column=1, value=f"WARNING: {analysis.date_mismatch_warning}")
            warning_cell.font = Font(bold=True, color="9A0000")
            warning_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            warning_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row_num}:G{row_num}')
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = self.BORDER
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            ws.row_dimensions[row_num].height = 35
            row_num += 1
        
        row_num += 1
        
        # Summary stats
        ws.cell(row=row_num, column=1, value="Total Items:").font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=len(analysis.items))
        
        ws.cell(row=row_num, column=4, value="Purchase Bills:").font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=analysis.purchase_bill_count)
        
        row_num += 1
        ws.cell(row=row_num, column=1, value="Items in Surplus:").font = Font(bold=True)
        surplus_count = len([i for i in analysis.items if i.surplus_deficit > 0])
        surplus_cell = ws.cell(row=row_num, column=2, value=surplus_count)
        surplus_cell.fill = self.SUCCESS_FILL
        
        ws.cell(row=row_num, column=4, value="Sales Bills:").font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=analysis.sales_bill_count)
        
        row_num += 1
        ws.cell(row=row_num, column=1, value="Items in Deficit:").font = Font(bold=True)
        deficit_count = len([i for i in analysis.items if i.surplus_deficit < 0])
        deficit_cell = ws.cell(row=row_num, column=2, value=deficit_count)
        if deficit_count > 0:
            deficit_cell.fill = self.ERROR_FILL
        
        # Value totals
        if analysis.total_purchase_value > 0 or analysis.total_sales_value > 0:
            ws.cell(row=row_num-2, column=6, value="Total Purchase:").font = Font(bold=True)
            pv_cell = ws.cell(row=row_num-2, column=7, value=analysis.total_purchase_value)
            pv_cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num-1, column=6, value="Total Sales:").font = Font(bold=True)
            sv_cell = ws.cell(row=row_num-1, column=7, value=analysis.total_sales_value)
            sv_cell.number_format = '#,##0.00'
        
        row_num += 2
        
        # Header row
        headers = [
            "Item Name", "Purch Qty", "Purch Value", "Sold Qty", "Sales Value", "Stock Balance", "Status"
        ]
        
        header_row = row_num
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Data rows with values
        row_num = header_row + 1
        for item in analysis.items:
            ws.cell(row=row_num, column=1, value=item.item_name.title()).border = self.BORDER
            
            # Purchased Qty
            pq_cell = ws.cell(row=row_num, column=2, value=item.purchased_qty)
            pq_cell.border = self.BORDER
            pq_cell.alignment = Alignment(horizontal='center')
            
            # Purchase Value
            pv_cell = ws.cell(row=row_num, column=3, value=item.purchased_value if item.purchased_value > 0 else "")
            pv_cell.border = self.BORDER
            pv_cell.alignment = Alignment(horizontal='right')
            pv_cell.number_format = '#,##0.00'
            
            # Sold Qty
            sq_cell = ws.cell(row=row_num, column=4, value=item.sold_qty)
            sq_cell.border = self.BORDER
            sq_cell.alignment = Alignment(horizontal='center')
            
            # Sales Value
            sv_cell = ws.cell(row=row_num, column=5, value=item.sold_value if item.sold_value > 0 else "")
            sv_cell.border = self.BORDER
            sv_cell.alignment = Alignment(horizontal='right')
            sv_cell.number_format = '#,##0.00'
            
            # Stock Balance (Surplus/Deficit) with color coding
            balance_cell = ws.cell(row=row_num, column=6, value=item.surplus_deficit)
            balance_cell.border = self.BORDER
            balance_cell.alignment = Alignment(horizontal='center')
            balance_cell.font = Font(bold=True)
            if item.surplus_deficit > 0:
                balance_cell.fill = self.SUCCESS_FILL
            elif item.surplus_deficit < 0:
                balance_cell.fill = self.ERROR_FILL
            
            # Status
            status_cell = ws.cell(row=row_num, column=7, value=item.status.value.title())
            status_cell.border = self.BORDER
            status_cell.alignment = Alignment(horizontal='center')
            if item.status.value == 'surplus':
                status_cell.fill = self.SUCCESS_FILL
            elif item.status.value == 'deficit':
                status_cell.fill = self.ERROR_FILL
            elif item.status.value == 'low_stock':
                status_cell.fill = self.WARNING_FILL
            
            row_num += 1
    
    def _create_bills_sheet(
        self, 
        wb: Workbook, 
        bills: List[dict], 
        sheet_name: str,
        bill_type: str
    ):
        """
        Create sheet with all bill line items including prices, discount, and GST.
        """
        ws = wb.create_sheet(sheet_name)
        
        # Column widths
        ws.column_dimensions['A'].width = 16  # Bill Number
        ws.column_dimensions['B'].width = 11  # Date
        ws.column_dimensions['C'].width = 22  # Vendor/Customer
        ws.column_dimensions['D'].width = 35  # Item Name
        ws.column_dimensions['E'].width = 8   # Qty
        ws.column_dimensions['F'].width = 12  # Rate
        ws.column_dimensions['G'].width = 10  # Disc. %
        ws.column_dimensions['H'].width = 14  # Amount
        ws.column_dimensions['I'].width = 12  # CGST
        ws.column_dimensions['J'].width = 12  # SGST
        ws.column_dimensions['K'].width = 14  # Total
        
        # Header
        party_type = "Vendor" if bill_type == "PURCHASE" else "Customer"
        headers = ["Bill Number", "Date", party_type, "Item Name", "Qty", "Rate", "Disc. %", "Amount", "CGST", "SGST", "Bill Total"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Data rows with prices, discount percentage, and GST
        row_num = 2
        grand_total_amount = 0.0
        grand_total_cgst = 0.0
        grand_total_sgst = 0.0
        grand_total_bill = 0.0
        
        for bill in bills:
            bill_number = bill.get('invoice_number', 'Unknown')
            date = bill.get('date', '')
            party = bill.get('vendor_name', '')
            line_items = bill.get('line_items', [])
            bill_cgst = bill.get('cgst', 0)
            bill_sgst = bill.get('sgst', 0)
            bill_total = bill.get('total', 0)
            
            # Track totals
            grand_total_cgst += bill_cgst
            grand_total_sgst += bill_sgst
            grand_total_bill += bill_total
            
            first_item = True
            
            # 1. Render Line Items (Products)
            for item in line_items:
                ws.cell(row=row_num, column=1, value=bill_number if first_item else "").border = self.BORDER
                ws.cell(row=row_num, column=2, value=date if first_item else "").border = self.BORDER
                ws.cell(row=row_num, column=3, value=party if first_item else "").border = self.BORDER
                ws.cell(row=row_num, column=4, value=item.item_name).border = self.BORDER
                
                qty_cell = ws.cell(row=row_num, column=5, value=item.quantity)
                qty_cell.alignment = Alignment(horizontal='center')
                qty_cell.border = self.BORDER
                
                rate_cell = ws.cell(row=row_num, column=6, value=item.rate if item.rate > 0 else "")
                rate_cell.alignment = Alignment(horizontal='right')
                rate_cell.border = self.BORDER
                rate_cell.number_format = '#,##0.00'
                
                # Discount percentage column
                disc_cell = ws.cell(row=row_num, column=7, value=f"{int(item.discount_percent)}%" if item.discount_percent > 0 else "")
                disc_cell.alignment = Alignment(horizontal='center')
                disc_cell.border = self.BORDER
                
                amount_cell = ws.cell(row=row_num, column=8, value=item.amount if item.amount > 0 else "")
                amount_cell.alignment = Alignment(horizontal='right')
                amount_cell.border = self.BORDER
                amount_cell.number_format = '#,##0.00'
                
                if item.amount > 0:
                    grand_total_amount += item.amount
                
                # Show GST and total only on first row of each bill
                if first_item:
                    cgst_cell = ws.cell(row=row_num, column=9, value=bill_cgst if bill_cgst > 0 else "")
                    cgst_cell.alignment = Alignment(horizontal='right')
                    cgst_cell.border = self.BORDER
                    cgst_cell.number_format = '#,##0.00'
                    
                    sgst_cell = ws.cell(row=row_num, column=10, value=bill_sgst if bill_sgst > 0 else "")
                    sgst_cell.alignment = Alignment(horizontal='right')
                    sgst_cell.border = self.BORDER
                    sgst_cell.number_format = '#,##0.00'
                    
                    total_cell = ws.cell(row=row_num, column=11, value=bill_total if bill_total > 0 else "")
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = self.BORDER
                    total_cell.number_format = '#,##0.00'
                    total_cell.font = Font(bold=True)
                else:
                    ws.cell(row=row_num, column=9).border = self.BORDER
                    ws.cell(row=row_num, column=10).border = self.BORDER
                    ws.cell(row=row_num, column=11).border = self.BORDER
                
                first_item = False
                row_num += 1
            
            # 2. Render Additional Charges (if any)
            charges = bill.get('additional_charges', [])
            for charge in charges:
                ws.cell(row=row_num, column=1, value=bill_number if first_item else "").border = self.BORDER
                ws.cell(row=row_num, column=2, value=date if first_item else "").border = self.BORDER
                ws.cell(row=row_num, column=3, value=party if first_item else "").border = self.BORDER
                
                # Show charge name with prefix
                charge_name_cell = ws.cell(row=row_num, column=4, value=f"Additional Charge: {charge.charge_name}")
                charge_name_cell.border = self.BORDER
                charge_name_cell.font = Font(italic=True, color="555555")
                
                # Empty quantity/rate/discount cells
                ws.cell(row=row_num, column=5).border = self.BORDER
                ws.cell(row=row_num, column=6).border = self.BORDER
                ws.cell(row=row_num, column=7).border = self.BORDER
                
                # Charge amount
                amount_cell = ws.cell(row=row_num, column=8, value=charge.amount)
                amount_cell.alignment = Alignment(horizontal='right')
                amount_cell.border = self.BORDER
                amount_cell.number_format = '#,##0.00'
                
                if charge.amount > 0:
                    grand_total_amount += charge.amount
                
                # Show GST/Total on first row if this is the ONLY item (no line items)
                if first_item:
                    cgst_cell = ws.cell(row=row_num, column=9, value=bill_cgst if bill_cgst > 0 else "")
                    cgst_cell.alignment = Alignment(horizontal='right')
                    cgst_cell.border = self.BORDER
                    cgst_cell.number_format = '#,##0.00'
                    
                    sgst_cell = ws.cell(row=row_num, column=10, value=bill_sgst if bill_sgst > 0 else "")
                    sgst_cell.alignment = Alignment(horizontal='right')
                    sgst_cell.border = self.BORDER
                    sgst_cell.number_format = '#,##0.00'
                    
                    total_cell = ws.cell(row=row_num, column=11, value=bill_total if bill_total > 0 else "")
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.border = self.BORDER
                    total_cell.number_format = '#,##0.00'
                    total_cell.font = Font(bold=True)
                else:
                    ws.cell(row=row_num, column=9).border = self.BORDER
                    ws.cell(row=row_num, column=10).border = self.BORDER
                    ws.cell(row=row_num, column=11).border = self.BORDER
                
                first_item = False
                row_num += 1
            
            # 3. IF NO LINE ITEMS & NO CHARGES FOUND
            # Print a single row so the bill doesn't disappear
            if first_item:
                ws.cell(row=row_num, column=1, value=bill_number).border = self.BORDER
                ws.cell(row=row_num, column=2, value=date).border = self.BORDER
                ws.cell(row=row_num, column=3, value=party).border = self.BORDER
                
                status_cell = ws.cell(row=row_num, column=4, value="[No Items Extracted]")
                status_cell.font = Font(italic=True, color="FF0000")
                status_cell.border = self.BORDER
                
                # Empty cells
                for col in range(5, 12):
                    ws.cell(row=row_num, column=col).border = self.BORDER
                
                # Can show total if it exists
                if bill_total > 0:
                     ws.cell(row=row_num, column=11, value=bill_total).border = self.BORDER
                
                row_num += 1
        
        # Grand Total row
        if grand_total_bill > 0:
            row_num += 1  # Empty row
            
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = self.BORDER
            
            # Empty cell for discount % column
            ws.cell(row=row_num, column=7).border = self.BORDER
            
            total_label = ws.cell(row=row_num, column=8, value="GRAND TOTAL:")
            total_label.font = Font(bold=True)
            total_label.border = self.BORDER
            
            cgst_total = ws.cell(row=row_num, column=9, value=grand_total_cgst)
            cgst_total.font = Font(bold=True)
            cgst_total.alignment = Alignment(horizontal='right')
            cgst_total.fill = self.SUCCESS_FILL
            cgst_total.border = self.BORDER
            cgst_total.number_format = '#,##0.00'
            
            sgst_total = ws.cell(row=row_num, column=10, value=grand_total_sgst)
            sgst_total.font = Font(bold=True)
            sgst_total.alignment = Alignment(horizontal='right')
            sgst_total.fill = self.SUCCESS_FILL
            sgst_total.border = self.BORDER
            sgst_total.number_format = '#,##0.00'
            
            bill_grand_total = ws.cell(row=row_num, column=11, value=grand_total_bill)
            bill_grand_total.font = Font(bold=True)
            bill_grand_total.alignment = Alignment(horizontal='right')
            bill_grand_total.fill = self.SUCCESS_FILL
            bill_grand_total.border = self.BORDER
            bill_grand_total.number_format = '#,##0.00'
        
        # If no data
        if row_num == 2:
            ws.cell(row=2, column=1, value=f"No {sheet_name.lower()} data")
            ws.merge_cells('A2:K2')
    
    def _create_insights_sheet(self, wb: Workbook, analysis):
        """
        Create sheet with stock management insights and recommendations.
        Redesigned with clean tabular format similar to Inventory Analysis Report.
        """
        ws = wb.create_sheet("Stock Insights")
        
        # Additional fill colors for insights
        TITLE_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Dark green
        SECTION_FILL = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")  # Blue
        METRIC_HEADER_FILL = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")  # Dark gray
        ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Light gray
        DEFICIT_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red
        LOW_STOCK_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
        TOP_SELLER_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
        
        # Column widths
        ws.column_dimensions['A'].width = 8    # S.No / Icon
        ws.column_dimensions['B'].width = 35   # Item Name / Metric
        ws.column_dimensions['C'].width = 15   # Value
        ws.column_dimensions['D'].width = 15   # Additional Info
        ws.column_dimensions['E'].width = 20   # Status/Action
        
        # ============ TITLE SECTION ============
        title_cell = ws.cell(row=1, column=1, value="STOCK MANAGEMENT INSIGHTS")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        for col in range(1, 6):
            ws.cell(row=1, column=col).border = self.BORDER
        
        # Generated timestamp
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=2, column=1).font = Font(italic=True, size=9)
        ws.merge_cells('A2:E2')
        
        # ============ KEY METRICS TABLE ============
        row_num = 4
        
        # Section header
        metrics_header = ws.cell(row=row_num, column=1, value="KEY METRICS")
        metrics_header.font = Font(bold=True, size=11, color="FFFFFF")
        metrics_header.fill = SECTION_FILL
        metrics_header.alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{row_num}:E{row_num}')
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = self.BORDER
        
        row_num += 1
        
        # Metrics table header
        metric_headers = ["#", "Metric", "Value", "Details", ""]
        for col, header in enumerate(metric_headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.fill = METRIC_HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        row_num += 1
        
        # Calculate metrics
        surplus_count = len([i for i in analysis.items if i.surplus_deficit > 0])
        deficit_count = len([i for i in analysis.items if i.surplus_deficit < 0])
        balanced_count = len([i for i in analysis.items if i.surplus_deficit == 0])
        low_stock_count = len(analysis.low_stock_items) if analysis.low_stock_items else 0
        
        # Metrics data
        metrics_data = [
            ("1", "Total Unique Items", str(len(analysis.items)), "Items tracked in inventory", None),
            ("2", "Purchase Bills Processed", str(analysis.purchase_bill_count), "Bills analyzed", None),
            ("3", "Sales Bills Processed", str(analysis.sales_bill_count), "Bills analyzed", None),
            ("4", "Items in Surplus", str(surplus_count), "Stock > Sales", self.SUCCESS_FILL),
            ("5", "Items in Deficit", str(deficit_count), "Sales > Stock (ALERT)", self.ERROR_FILL if deficit_count > 0 else None),
            ("6", "Items Balanced", str(balanced_count), "Stock = Sales", None),
            ("7", "Low Stock Items", str(low_stock_count), "Need reordering", self.WARNING_FILL if low_stock_count > 0 else None),
        ]
        
        for idx, (num, metric, value, details, fill) in enumerate(metrics_data):
            ws.cell(row=row_num, column=1, value=num).border = self.BORDER
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_num, column=2, value=metric).border = self.BORDER
            
            value_cell = ws.cell(row=row_num, column=3, value=value)
            value_cell.border = self.BORDER
            value_cell.alignment = Alignment(horizontal='center')
            value_cell.font = Font(bold=True)
            if fill:
                value_cell.fill = fill
            
            ws.cell(row=row_num, column=4, value=details).border = self.BORDER
            ws.cell(row=row_num, column=5).border = self.BORDER
            
            # Alternate row coloring
            if idx % 2 == 1:
                for col in range(1, 6):
                    if not fill or col != 3:
                        ws.cell(row=row_num, column=col).fill = ALT_ROW_FILL
            
            row_num += 1
        
        row_num += 1
        
        # ============ KEY INSIGHTS SECTION ============
        if analysis.insights:
            insights_header = ws.cell(row=row_num, column=1, value="KEY INSIGHTS")
            insights_header.font = Font(bold=True, size=11, color="FFFFFF")
            insights_header.fill = SECTION_FILL
            insights_header.alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row_num}:E{row_num}')
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = self.BORDER
            row_num += 1
            
            # Table header for insights
            insight_headers = ["#", "Type", "Insight Details", "", ""]
            for col, header in enumerate(insight_headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.fill = METRIC_HEADER_FILL
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.BORDER
            row_num += 1
            
            for idx, insight in enumerate(analysis.insights, 1):
                # Determine insight type and color
                if "[CRITICAL]" in insight:
                    insight_type = "CRITICAL"
                    type_fill = self.ERROR_FILL
                    insight_text = insight.replace("[CRITICAL] ", "")
                elif "[ALERT]" in insight:
                    insight_type = "ALERT"
                    type_fill = self.WARNING_FILL
                    insight_text = insight.replace("[ALERT] ", "")
                elif "[GOOD]" in insight:
                    insight_type = "GOOD"
                    type_fill = self.SUCCESS_FILL
                    insight_text = insight.replace("[GOOD] ", "")
                elif "[TOP]" in insight:
                    insight_type = "TOP"
                    type_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")  # Light blue
                    insight_text = insight.replace("[TOP] ", "")
                elif "[SUMMARY]" in insight:
                    insight_type = "SUMMARY"
                    type_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Gray
                    insight_text = insight.replace("[SUMMARY] ", "")
                else:
                    insight_type = "INFO"
                    type_fill = ALT_ROW_FILL
                    insight_text = insight
                
                # Row number
                num_cell = ws.cell(row=row_num, column=1, value=str(idx))
                num_cell.border = self.BORDER
                num_cell.alignment = Alignment(horizontal='center')
                
                # Type badge
                type_cell = ws.cell(row=row_num, column=2, value=insight_type)
                type_cell.border = self.BORDER
                type_cell.fill = type_fill
                type_cell.font = Font(bold=True, size=9)
                type_cell.alignment = Alignment(horizontal='center')
                
                # Insight text
                insight_cell = ws.cell(row=row_num, column=3, value=insight_text)
                insight_cell.border = self.BORDER
                ws.merge_cells(f'C{row_num}:E{row_num}')
                for col in range(3, 6):
                    ws.cell(row=row_num, column=col).border = self.BORDER
                
                row_num += 1
            
            row_num += 1

        
        # ============ DEFICIT ITEMS TABLE ============
        if analysis.deficit_items:
            deficit_header = ws.cell(row=row_num, column=1, value="DEFICIT ITEMS (Need Procurement)")
            deficit_header.font = Font(bold=True, size=11, color="FFFFFF")
            deficit_header.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")  # Red
            deficit_header.alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row_num}:E{row_num}')
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = self.BORDER
            row_num += 1
            
            # Table header
            def_headers = ["#", "Item Name", "Status", "Action Required", "Priority"]
            for col, header in enumerate(def_headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.fill = METRIC_HEADER_FILL
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.BORDER
            row_num += 1
            
            for idx, item in enumerate(analysis.deficit_items[:10], 1):
                ws.cell(row=row_num, column=1, value=str(idx)).border = self.BORDER
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
                
                name_cell = ws.cell(row=row_num, column=2, value=item.title())
                name_cell.border = self.BORDER
                
                status_cell = ws.cell(row=row_num, column=3, value="Deficit")
                status_cell.border = self.BORDER
                status_cell.fill = DEFICIT_FILL
                status_cell.alignment = Alignment(horizontal='center')
                
                action_cell = ws.cell(row=row_num, column=4, value="Procure Stock")
                action_cell.border = self.BORDER
                action_cell.alignment = Alignment(horizontal='center')
                
                priority_cell = ws.cell(row=row_num, column=5, value="HIGH" if idx <= 3 else "MEDIUM")
                priority_cell.border = self.BORDER
                priority_cell.alignment = Alignment(horizontal='center')
                priority_cell.font = Font(bold=True)
                if idx <= 3:
                    priority_cell.fill = self.ERROR_FILL
                else:
                    priority_cell.fill = self.WARNING_FILL
                
                row_num += 1
            
            row_num += 1
        
        # ============ LOW STOCK ITEMS TABLE ============
        if analysis.low_stock_items:
            low_stock_header = ws.cell(row=row_num, column=1, value="LOW STOCK ITEMS (Consider Reordering)")
            low_stock_header.font = Font(bold=True, size=11, color="FFFFFF")
            low_stock_header.fill = PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid")  # Amber
            low_stock_header.alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row_num}:E{row_num}')
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = self.BORDER
            row_num += 1
            
            # Table header
            low_headers = ["#", "Item Name", "Status", "Recommendation", "Priority"]
            for col, header in enumerate(low_headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.fill = METRIC_HEADER_FILL
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.BORDER
            row_num += 1
            
            for idx, item in enumerate(analysis.low_stock_items[:10], 1):
                ws.cell(row=row_num, column=1, value=str(idx)).border = self.BORDER
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
                
                name_cell = ws.cell(row=row_num, column=2, value=item.title())
                name_cell.border = self.BORDER
                
                status_cell = ws.cell(row=row_num, column=3, value="Low Stock")
                status_cell.border = self.BORDER
                status_cell.fill = LOW_STOCK_FILL
                status_cell.alignment = Alignment(horizontal='center')
                
                rec_cell = ws.cell(row=row_num, column=4, value="Reorder Soon")
                rec_cell.border = self.BORDER
                rec_cell.alignment = Alignment(horizontal='center')
                
                priority_cell = ws.cell(row=row_num, column=5, value="MEDIUM")
                priority_cell.border = self.BORDER
                priority_cell.alignment = Alignment(horizontal='center')
                priority_cell.fill = self.WARNING_FILL
                
                row_num += 1
            
            row_num += 1
        
        # ============ TOP SELLING ITEMS TABLE ============
        if analysis.top_selling_items:
            top_header = ws.cell(row=row_num, column=1, value="TOP SELLING ITEMS")
            top_header.font = Font(bold=True, size=11, color="FFFFFF")
            top_header.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Green
            top_header.alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A{row_num}:E{row_num}')
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = self.BORDER
            row_num += 1
            
            # Table header
            top_headers = ["Rank", "Item Name", "Performance", "Stock Status", ""]
            for col, header in enumerate(top_headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.fill = METRIC_HEADER_FILL
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.BORDER
            row_num += 1
            
            # Get item data for stock status check
            item_data = {i.item_name.lower(): i for i in analysis.items}
            
            for idx, item in enumerate(analysis.top_selling_items[:5], 1):
                # Rank indicator
                rank_text = f"#{idx}"
                rank_cell = ws.cell(row=row_num, column=1, value=rank_text)
                rank_cell.border = self.BORDER
                rank_cell.alignment = Alignment(horizontal='center')
                rank_cell.font = Font(size=12 if idx <= 3 else 10)
                
                name_cell = ws.cell(row=row_num, column=2, value=item.title())
                name_cell.border = self.BORDER
                name_cell.font = Font(bold=True if idx == 1 else False)
                
                perf_cell = ws.cell(row=row_num, column=3, value="Top Seller")
                perf_cell.border = self.BORDER
                perf_cell.fill = TOP_SELLER_FILL
                perf_cell.alignment = Alignment(horizontal='center')
                
                # Check stock status
                item_info = item_data.get(item.lower())
                stock_status = "Unknown"
                if item_info:
                    if item_info.surplus_deficit > 0:
                        stock_status = "In Stock"
                    elif item_info.surplus_deficit < 0:
                        stock_status = "REORDER!"
                    else:
                        stock_status = "Balanced"
                
                stock_cell = ws.cell(row=row_num, column=4, value=stock_status)
                stock_cell.border = self.BORDER
                stock_cell.alignment = Alignment(horizontal='center')
                if "REORDER" in stock_status:
                    stock_cell.fill = self.ERROR_FILL
                    stock_cell.font = Font(bold=True)
                elif "In Stock" in stock_status:
                    stock_cell.fill = self.SUCCESS_FILL
                
                ws.cell(row=row_num, column=5).border = self.BORDER
                
                row_num += 1
        
        # ============ FOOTER ============
        row_num += 2
        footer_cell = ws.cell(row=row_num, column=1, value="Report generated by DATTU_BILL Stock Management System")
        footer_cell.font = Font(italic=True, size=9, color="666666")
        ws.merge_cells(f'A{row_num}:E{row_num}')
